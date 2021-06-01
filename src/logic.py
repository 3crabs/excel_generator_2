import datetime
from shutil import copyfile

import openpyxl
from openpyxl import load_workbook


def read_file(file_path: str) -> list:
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    return sheet.rows


m = {
    '1': [{'s': 'стр.1_2', 'c': 'AK9'}, {'s': 'стр.1_2', 'c': 'CM9'}, {'s': 'стр.1_2', 'c': 'BC46'}],
    '2': [{'s': 'стр.1_2', 'c': 'BN46'}],
    '3': [{'s': 'стр.1_2', 'c': 'BD16'}, {'s': 'стр.1_2', 'c': 'H48'}],
    '4': [{'s': 'стр.1_2', 'c': 'BD17'}, {'s': 'стр.1_2', 'c': 'BD59'}],
    '5': [],
    '6': [{'s': 'стр.1_2', 'c': 'AO30'}, {'s': 'стр.1_2', 'c': 'AM69'}, {'s': 'стр.1_2', 'c': 'CO69'}],
    '7': [{'s': 'стр.1_2', 'c': 'B30'}, {'s': 'стр.1_2', 'c': 'AD69'}, {'s': 'стр.1_2', 'c': 'CF69'}],
    '8': [{'s': 'стр.1_2', 'c': 'B33'}, {'s': 'стр.1_2', 'c': 'B69'}, {'s': 'стр.1_2', 'c': 'BD69'}],
    '9': [{'s': 'стр.1_2', 'c': 'B38'}, {'s': 'стр.1_2', 'c': 'B55'}],
    '10': [{'s': 'стр.1_2', 'c': 'BH18'}, {'s': 'стр.1_2', 'c': 'CR48'}],
    '11': [{'s': 'стр.1_2', 'c': 'BH19'}],
    '12': [{'s': 'стр.1_2', 'c': 'G9'}, {'s': 'стр.1_2', 'c': 'B63'}, {'s': 'стр.1_2', 'c': 'BD63'},
           {'s': 'стр.1_2', 'c': 'B65'}, {'s': 'стр.1_2', 'c': 'AD65'}, {'s': 'стр.1_2', 'c': 'BD65'},
           {'s': 'стр.1_2', 'c': 'CF65'}],
    '13': [{'s': 'стр.1_2', 'c': 'BI9'}, {'s': 'стр.1_2', 'c': 'B91'}],
    '14': [{'s': 'стр.1_2', 'c': 'I75'}, {'s': 'стр.1_2', 'c': 'BK75'}, {'s': 'стр.1_2', 'c': 'U101'}],
    '15': [{'s': 'стр.1_2', 'c': 'K104'}],
    '16': [{'s': 'стр.1_2', 'c': 'BD104'}],
}


def make_invoice_file(invoice: list, template_file_path: str, save_dir_path: str):
    if invoice[0].value is None:
        return
    new_file_path = save_dir_path + "/" + invoice[0].value + ".xlsx"
    copyfile(template_file_path, new_file_path)
    wb = openpyxl.load_workbook(filename=new_file_path)

    i = 1
    for cell in invoice:
        for place in m[str(i)]:
            s = cell.value
            if datetime.datetime == type(s):
                s = s.strftime("%d.%m.%Y")
            s = str(s)
            if i == 9:
                s = s.replace('.', ',')
            wb[place['s']][place['c']] = s
        i += 1

    wb.save(new_file_path)
