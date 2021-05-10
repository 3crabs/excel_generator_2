from openpyxl import load_workbook

from src.consts import path_template_file, path_out_dir
from src.logic import read_file, make_invoice_file


def test_1():
    wb = load_workbook("../resources/template.xlsx")
    for sheet_name in wb.sheetnames:
        print(sheet_name)


def test_2():
    read_book = load_workbook("/home/vladimir/Проекты/Генератор файлов/Задание/Копия реестр 2021.xlsx")
    for sheet_name in read_book.sheetnames:
        print(sheet_name)


def test_3():
    invoices = read_file("/home/vladimir/Проекты/Генератор файлов/Задание/Копия реестр 2021.xlsx")
    for invoice in list(invoices)[2:]:
        make_invoice_file(invoice, path_template_file, path_out_dir)


if __name__ == '__main__':
    test_1()
    print()
    test_2()
    print()
    test_3()
